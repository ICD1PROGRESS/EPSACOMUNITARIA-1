from sqlmodel import Session, select, func
from typing import List, Optional
from .models import EPSA, Usuario, ConfiguracionEPSA, Periodo, Tarifa, Lectura, Pago, Deuda, Plan, Suscripcion
import openpyxl
from pathlib import Path
from datetime import date, datetime, timedelta

# ---------- EPSA ----------
def create_epsa(session: Session, nombre: str, ciudad: str) -> EPSA:
    epsa = EPSA(nombre=nombre, ciudad=ciudad)
    session.add(epsa)
    session.commit()
    session.refresh(epsa)
    return epsa

def get_all_epsas(session: Session) -> List[EPSA]:
    return session.exec(select(EPSA)).all()

# ---------- Usuarios ----------
def create_usuario(session: Session, epsa_id: int, codigo: str, nombre: str, ci: str, zona: str, nro_medidor: str, categoria: str = "RESIDENCIAL") -> Usuario:
    usuario = Usuario(
        epsa_id=epsa_id,
        codigo=codigo,
        nombre=nombre,
        ci=ci,
        zona=zona,
        nro_medidor=nro_medidor,
        categoria=categoria,
        saldo_actual=0.0
    )
    session.add(usuario)
    session.commit()
    session.refresh(usuario)
    return usuario

def get_usuarios_by_epsa(session: Session, epsa_id: int) -> List[Usuario]:
    return session.exec(select(Usuario).where(Usuario.epsa_id == epsa_id)).all()

def get_usuario_by_codigo(session: Session, epsa_id: int, codigo: str) -> Optional[Usuario]:
    return session.exec(select(Usuario).where(Usuario.epsa_id == epsa_id, Usuario.codigo == codigo)).first()

def update_usuario(session: Session, usuario_id: int, **kwargs):
    usuario = session.get(Usuario, usuario_id)
    if usuario:
        for key, value in kwargs.items():
            setattr(usuario, key, value)
        session.commit()
        session.refresh(usuario)
    return usuario

def delete_usuario(session: Session, usuario_id: int):
    usuario = session.get(Usuario, usuario_id)
    if usuario:
        session.delete(usuario)
        session.commit()

def importar_usuarios_desde_excel(session: Session, epsa_id: int, file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]
    col_idx = {k: None for k in ['codigo', 'nombre', 'ci', 'zona', 'nro_medidor', 'categoria']}
    for idx, h in enumerate(headers, start=1):
        if h in col_idx:
            col_idx[h] = idx
    required = ['codigo', 'nombre', 'ci']
    missing = [r for r in required if col_idx[r] is None]
    if missing:
        return {"total": 0, "insertados": 0, "errores": [f"Faltan columnas: {', '.join(missing)}"]}
    total = 0
    insertados = 0
    errores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        total += 1
        try:
            codigo = str(row[col_idx['codigo']-1]).strip()
            nombre = str(row[col_idx['nombre']-1]).strip()
            ci = str(row[col_idx['ci']-1]).strip()
            zona = str(row[col_idx['zona']-1]) if col_idx['zona'] else ''
            nro_medidor = str(row[col_idx['nro_medidor']-1]) if col_idx['nro_medidor'] else ''
            #categoria = str(row[col_idx['categoria']-1]) if col_idx['categoria'] else 'RESIDENCIAL'
            #categoria = categoria.upper()
            categoria_raw = str(row[col_idx['categoria']-1]) if col_idx['categoria'] and row[col_idx['categoria']-1] else ''
            categoria = categoria_raw.upper().strip()
            if categoria not in ['RESIDENCIAL', 'COMERCIAL', 'INDUSTRIAL', 'PÚBLICO']:
                categoria = 'RESIDENCIAL'

            existente = session.exec(select(Usuario).where(Usuario.epsa_id == epsa_id, Usuario.codigo == codigo)).first()
            if existente:
                errores.append(f"Fila {total+1}: Código {codigo} ya existe")
                continue
            nuevo = Usuario(epsa_id=epsa_id, codigo=codigo, nombre=nombre, ci=ci, zona=zona, nro_medidor=nro_medidor, categoria=categoria)
            session.add(nuevo)
            insertados += 1
        except Exception as e:
            errores.append(f"Fila {total+1}: {str(e)}")
    session.commit()
    return {"total": total, "insertados": insertados, "errores": errores}

# ---------- Tarifas ----------
def create_tarifa(session: Session, epsa_id: int, rango_inicio: float, precio_unitario: float, rango_fin: Optional[float] = None, cargo_fijo: float = 0.0, orden: int = 0):
    tarifa = Tarifa(epsa_id=epsa_id, rango_inicio=rango_inicio, rango_fin=rango_fin, precio_unitario=precio_unitario, cargo_fijo=cargo_fijo, orden=orden)
    session.add(tarifa)
    session.commit()
    session.refresh(tarifa)
    return tarifa

def get_tarifas_by_epsa(session: Session, epsa_id: int) -> List[Tarifa]:
    return session.exec(select(Tarifa).where(Tarifa.epsa_id == epsa_id).order_by(Tarifa.orden, Tarifa.rango_inicio)).all()

def delete_tarifa(session: Session, tarifa_id: int):
    tarifa = session.get(Tarifa, tarifa_id)
    if tarifa:
        session.delete(tarifa)
        session.commit()

def actualizar_saldo_usuario(session: Session, usuario_id: int, monto_cargo: float):
    usuario = session.get(Usuario, usuario_id)
    if usuario:
        usuario.saldo_actual += monto_cargo
        session.add(usuario)
        session.commit()

def get_ultima_lectura(session: Session, usuario_id: int) -> Optional[Lectura]:
    return session.exec(
        select(Lectura).where(Lectura.usuario_id == usuario_id).order_by(Lectura.fecha_toma.desc())
    ).first()

def get_usuario_by_id(session: Session, usuario_id: int) -> Optional[Usuario]:
    """Obtiene un usuario por su ID."""
    return session.get(Usuario, usuario_id)

# ---------- Funciones auxiliares para Caja ----------

def get_periodo_activo(session: Session, epsa_id: int):
    """Obtiene el periodo activo de una EPSA."""
    return session.exec(
        select(Periodo).where(Periodo.epsa_id == epsa_id, Periodo.activo == True)
    ).first()

def get_pago_by_usuario_periodo(session: Session, usuario_id: int, periodo: str):
    """Verifica si un usuario ya pagó un periodo específico."""
    return session.exec(
        select(Pago).where(Pago.usuario_id == usuario_id, Pago.periodo == periodo)
    ).first()

def get_periodo_activo_nombre(session: Session, epsa_id: int) -> str:
    """Obtiene el nombre del periodo activo. Fallback a última lectura."""
    periodo_obj = session.exec(
        select(Periodo).where(Periodo.epsa_id == epsa_id, Periodo.activo == True)
    ).first()
    if periodo_obj:
        return periodo_obj.nombre
    
    ult_lectura = session.exec(
        select(Lectura).where(Lectura.epsa_id == epsa_id).order_by(Lectura.periodo.desc())
    ).first()
    if ult_lectura:
        return ult_lectura.periodo
    return "SIN PERIODO"

# ===================================================================
# FUNCIONES PARA CIERRE DE PERIODO (actualiza.frm)
# ===================================================================

def get_lecturas_periodo_detalle(session: Session, epsa_id: int, periodo: str):
    """
    Obtiene lecturas del periodo con datos del usuario, estado de pago y deudas anteriores.
    Equivalente a leer la hoja 'base' del macro con todas sus columnas.
    """
    from .models import Lectura, Usuario, Pago, Deuda
    
    lecturas = session.exec(
        select(Lectura, Usuario)
        .join(Usuario, Lectura.usuario_id == Usuario.id)
        .where(Lectura.epsa_id == epsa_id, Lectura.periodo == periodo)
    ).all()
    
    resultado = []
    for lect, usuario in lecturas:
        # ¿Pagó este periodo? (equivalente a columna O=1)
        pago = session.exec(
            select(Pago).where(Pago.usuario_id == usuario.id, Pago.periodo == periodo)
        ).first()
        
        # Deudas pendientes anteriores (equivalente a columna K antes del cierre)
        deudas_ant = session.exec(
            select(Deuda).where(Deuda.usuario_id == usuario.id, Deuda.estado == "PENDIENTE")
        ).all()
        total_deuda_ant = sum(d.monto for d in deudas_ant)
        
        # Total a pagar = importe periodo + deuda anterior (equivalente a columna N)
        total = lect.importe_calculado + total_deuda_ant
        
        resultado.append({
            "usuario_id": usuario.id,
            "codigo": usuario.codigo,
            "nombre": usuario.nombre,
            "lectura_anterior": lect.lectura_anterior,
            "lectura_actual": lect.lectura_actual,
            "consumo_m3": lect.consumo_m3,
            "importe_calculado": lect.importe_calculado,
            "deuda_anterior": total_deuda_ant,
            "total": total,
            "pagado": pago is not None,
            "lectura_id": lect.id
        })
    
    return resultado

def ejecutar_cierre_periodo(session: Session, epsa_id: int, periodo_nombre: str,
                            nuevo_periodo_nombre: str, fecha_cierre=None):
    
    if fecha_cierre is None:
        fecha_cierre = date.today()
    
    resumen = {
        "lecturas_procesadas": 0,
        "deudas_generadas": 0,
        "deudas_pagadas": 0,
        "usuarios_sin_lectura": 0,
        "errores": []
    }
    
    # 1. OBTENER LECTURAS DEL PERIODO (equivalente a recorrer hoja "base")
    lecturas = session.exec(
        select(Lectura).where(
            Lectura.epsa_id == epsa_id,
            Lectura.periodo == periodo_nombre
        )
    ).all()
    
    resumen["lecturas_procesadas"] = len(lecturas)
    
    for lect in lecturas:
        usuario = session.get(Usuario, lect.usuario_id)
        if not usuario:
            resumen["errores"].append(f"Usuario ID {lect.usuario_id} no encontrado")
            continue
        
        # Verificar si pagó (columna O=1 en macro)
        pago = session.exec(
            select(Pago).where(
                Pago.usuario_id == usuario.id,
                Pago.periodo == periodo_nombre
            )
        ).first()
        
        if pago:
            # Pagó: en el macro se generaba deuda como "Pagado" y luego se eliminaba de la hoja.
            # En la app, no generamos deuda (el pago ya quedó en tabla Pago).
            resumen["deudas_pagadas"] += 1
        else:
            # NO pagó: generar deuda en tabla Deuda (equivalente a copiar en hoja "deuda")
            max_id = session.exec(select(func.max(Deuda.id))).first() or 0
            nro_deuda = f"DEU-{max_id + 1:04d}"
            
            deuda = Deuda(
                epsa_id=epsa_id,
                usuario_id=usuario.id,
                nro_deuda=nro_deuda,
                periodo=periodo_nombre,
                consumo_m3=lect.consumo_m3,
                monto=lect.importe_calculado,  # importe + reconexión si aplica
                estado="PENDIENTE",
                fecha_vencimiento=fecha_cierre + timedelta(days=30)
            )
            session.add(deuda)
            
            # Actualizar saldo acumulado del usuario (equivalente a columna K en macro)
            usuario.saldo_actual += lect.importe_calculado
            session.add(usuario)
            resumen["deudas_generadas"] += 1
            
    # 2. CERRAR PERIODO ACTUAL
    periodo_actual = session.exec(
        select(Periodo).where(
            Periodo.epsa_id == epsa_id,
            Periodo.nombre == periodo_nombre
        )
    ).first()
    
    if periodo_actual:
        periodo_actual.activo = False
        periodo_actual.cerrado = True
        periodo_actual.fecha_fin = fecha_cierre
        session.add(periodo_actual)
    else:
        periodo_actual = Periodo(
            epsa_id=epsa_id,
            nombre=periodo_nombre,
            fecha_inicio=fecha_cierre,
            fecha_fin=fecha_cierre,
            activo=False,
            cerrado=True
        )
        session.add(periodo_actual)
    
    # 3. CREAR NUEVO PERIODO (equivalente a defper_Click + Lperiodo)
    nuevo_periodo = Periodo(
        epsa_id=epsa_id,
        nombre=nuevo_periodo_nombre,
        fecha_inicio=fecha_cierre,
        fecha_fin=fecha_cierre,
        activo=True,
        cerrado=False
    )
    session.add(nuevo_periodo)
    
    session.commit()
    resumen["nuevo_periodo"] = nuevo_periodo_nombre
    
    return resumen

def get_ultima_lectura_periodo_cerrado(session: Session, usuario_id: int, periodo_activo: str):
    """
    Obtiene la última lectura del usuario EXCLUYENDO el periodo activo.
    Equivalente a obtener la 'lectura anterior' después del cierre en el macro.
    """
    return session.exec(
        select(Lectura)
        .where(Lectura.usuario_id == usuario_id)
        .where(Lectura.periodo != periodo_activo)  # Excluir periodo activo
        .order_by(Lectura.fecha_toma.desc())
    ).first()

#CODIGO TEMPORAL
def cerrar_periodo(session: Session, epsa_id: int, periodo_actual: str, fecha_cierre: date = None):
    """
    Ejecuta el cierre de período:
    - Calcula consumo e importe de lecturas actuales
    - Actualiza saldo_actual de usuarios
    - Mueve lecturas actuales a anterior (o las marca como históricas)
    - Limpia lecturas actuales para nuevo período
    - Cierra el período actual
    """
    from .models import Lectura, Usuario
    from src.business.tarifa_service import calcular_importe_consumo
    from src.database.crud import get_tarifas_by_epsa
    
    if fecha_cierre is None:
        fecha_cierre = date.today()
    
    # Obtener todas las lecturas del período actual que tengan lectura_actual > 0
    lecturas = session.exec(
        select(Lectura).where(Lectura.epsa_id == epsa_id, Lectura.lectura_actual > 0, Lectura.periodo == periodo_actual)
    ).all()
    
    tarifas = get_tarifas_by_epsa(session, epsa_id)
    
    for lect in lecturas:
        consumo = lect.lectura_actual - lect.lectura_anterior
        if consumo < 0:
            consumo = 0  # O podrías manejar error
        lect.consumo_m3 = round(consumo, 2)
        if tarifas:
            importe = calcular_importe_consumo(consumo, tarifas)
        else:
            importe = 0
        lect.importe_calculado = importe
        # Actualizar saldo del usuario
        usuario = session.get(Usuario, lect.usuario_id)
        if usuario:
            usuario.saldo_actual += importe
            session.add(usuario)
        # Aquí podrías marcar la lectura como "cerrada" o simplemente no borrarla
        # Para el próximo período, se creará un nuevo registro de lectura.
    session.commit()
    
    # Marcar el período como cerrado
    periodo_obj = session.exec(select(Periodo).where(Periodo.epsa_id == epsa_id, Periodo.activo == True)).first()
    if periodo_obj:
        periodo_obj.activo = False
        periodo_obj.cerrado = True
        periodo_obj.fecha_fin = fecha_cierre
        session.add(periodo_obj)
    session.commit()
    
    return len(lecturas)
